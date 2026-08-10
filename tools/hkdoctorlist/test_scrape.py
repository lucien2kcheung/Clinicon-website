#!/usr/bin/env python3
"""End-to-end check of scrape.py against a local mock of the site's structure.

Serves listing pages at /type/<slug>/<n>/ and profile pages at /doctor/<id>/,
in three different markup styles (JSON-LD, definition list, table), then runs
the real crawler against it and asserts the workbook comes out correct.

    python test_scrape.py
"""

from __future__ import annotations

import http.server
import socketserver
import sys
import tempfile
import threading
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import scrape  # noqa: E402

PER_PAGE = 5
CATEGORY_SIZES = {"dentists": 12, "chinese-medicine": 8, "doctors": 14, "physiotherapists": 6}
DISTRICTS = ["中西區", "灣仔區", "東區"]
SPECIALTIES = ["牙科普通科", "內科", "骨科"]


def profile_id(slug: str, index: int) -> str:
    return f"{slug}-{index}"


def listing_html(slug: str, page: int) -> str:
    total = CATEGORY_SIZES[slug]
    pages = (total + PER_PAGE - 1) // PER_PAGE
    start = (page - 1) * PER_PAGE
    rows = "".join(
        f'<li><a href="/doctor/{profile_id(slug, i)}/">醫師 {i}</a></li>'
        for i in range(start, min(start + PER_PAGE, total))
    )
    pager = "".join(f'<a href="/type/{slug}/{p}/">{p}</a> ' for p in range(1, pages + 1))
    nav = "".join(f'<a href="/type/{s}/">{s}</a> ' for s in CATEGORY_SIZES)
    return f"""<html><head><title>{slug} 名單</title></head><body>
    <nav>{nav}<a href="/about/">關於</a></nav>
    <ul>{rows}</ul><div class="pagination">{pager}</div></body></html>"""


def type_index_html() -> str:
    labels = {"dentists": "牙醫", "chinese-medicine": "中醫師",
              "doctors": "醫生", "physiotherapists": "物理治療師"}
    links = "".join(f'<a href="/type/{s}/">{l}名單</a> ' for s, l in labels.items())
    return f"<html><body><h1>按類別</h1>{links}</body></html>"


def profile_html(slug: str, index: int) -> str:
    """Rotate through three markup styles so all parser paths get exercised."""
    name = f"陳大文 {slug} {index}"
    phone = f"2{index:03d} {1000 + index}"
    fax = f"3{index:03d} {2000 + index}"
    email = f"{slug}{index}@example.com"
    address = f"香港{DISTRICTS[index % 3]}示範道 {index} 號 {index}樓"
    district = DISTRICTS[index % 3]
    specialty = SPECIALTIES[index % 3]
    crumbs = (f'<a href="/district/d{index % 3}/">{district}</a>'
              f'<a href="/specialist/s{index % 3}/">{specialty}</a>')
    style = index % 3

    if style == 0:
        body = f"""<script type="application/ld+json">{{
          "@context":"https://schema.org","@type":"Physician",
          "name":"{name}","telephone":"{phone}","faxNumber":"{fax}",
          "email":"{email}","medicalSpecialty":"{specialty}",
          "address":{{"@type":"PostalAddress","streetAddress":"{address}",
                      "addressLocality":"{district}"}}
        }}</script><h1>{name}</h1>{crumbs}"""
    elif style == 1:
        body = f"""<h1>{name}</h1>{crumbs}
        <dl><dt>診所名稱</dt><dd>{slug} 醫務中心</dd>
            <dt>地址</dt><dd>{address}</dd>
            <dt>電話</dt><dd>{phone}</dd>
            <dt>傳真</dt><dd>{fax}</dd>
            <dt>電郵</dt><dd>{email}</dd>
            <dt>地區</dt><dd>{district}</dd>
            <dt>專科</dt><dd>{specialty}</dd>
            <dt>應診時間</dt><dd>星期一至五 09:00-18:00</dd></dl>"""
    else:
        body = f"""<h1>{name}</h1>{crumbs}
        <table>
          <tr><th>地址</th><td>{address}</td></tr>
          <tr><th>電話</th><td>{phone}</td></tr>
          <tr><th>傳真</th><td>{fax}</td></tr>
          <tr><th>註冊編號</th><td>M{10000 + index}</td></tr>
          <tr><th>地區</th><td>{district}</td></tr>
          <tr><th>專科</th><td>{specialty}</td></tr>
        </table>
        <p>電郵：{email}</p><a href="mailto:{email}">聯絡</a>"""

    return f"<html><head><title>{name} | hkdoctorlist</title></head><body>{body}</body></html>"


class Handler(http.server.BaseHTTPRequestHandler):
    def log_message(self, *args):  # silence per-request logging
        pass

    def do_GET(self):
        path = self.path.split("?")[0]
        parts = [p for p in path.split("/") if p]
        html = None

        if path == "/robots.txt":
            self.send_response(200)
            self.send_header("Content-Type", "text/plain")
            self.end_headers()
            self.wfile.write(b"User-agent: *\nDisallow: /admin/\n")
            return

        if parts == ["type"]:
            html = type_index_html()
        elif len(parts) >= 2 and parts[0] == "type" and parts[1] in CATEGORY_SIZES:
            slug = parts[1]
            page = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 1
            pages = (CATEGORY_SIZES[slug] + PER_PAGE - 1) // PER_PAGE
            if 1 <= page <= pages:
                html = listing_html(slug, page)
        elif len(parts) == 2 and parts[0] == "doctor":
            slug, _, idx = parts[1].rpartition("-")
            if slug in CATEGORY_SIZES and idx.isdigit():
                html = profile_html(slug, int(idx))

        if html is None:
            self.send_error(404)
            return
        payload = html.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)


# A faithful reduction of a real hkdoctorlist detail page, kept as a regression
# guard for the things the live site does that the generic parser got wrong:
# a site-wide Organization JSON-LD block ahead of the practitioner's own node,
# labels sitting in <h2> headings rather than <dl>/<table> cells, an office
# hours heading nested one level deeper than its value, an empty Languages
# block immediately before the next column, and emails the site itself mangles
# by replacing "tel" with "電話".
LIVE_MARKUP = """<!DOCTYPE html><html lang="zh-HK"><head>
<title>Dr. Jain Sandeep — 牙科普通科, 中西區 | hkdoctorlist.com.hk</title>
<script type="application/ld+json">{"@context":"https://schema.org","@type":"Organization",
"name":"香港醫生資料庫 hkdoctorlist.com.hk","url":"https://www.hkdoctorlist.com.hk"}</script>
<script type="application/ld+json">{"@context":"https://schema.org","@type":"BreadcrumbList",
"itemListElement":[{"@type":"ListItem","position":1,"name":"首頁 Home"}]}</script>
<script type="application/ld+json">{"@context":"https://schema.org","@type":"Dentist",
"name":"Dr. Jain Sandeep","medicalSpecialty":"牙科普通科","address":{"@type":"PostalAddress",
"streetAddress":"Room 1914, Melbourne Plaza, 33 Queen's Road Central, Hong Kong",
"addressLocality":"Central & Western","addressRegion":"Hong Kong"},
"telephone":"25222099","email":"drjain@dies電話andpartners.com"}</script>
</head><body>
<div><div class="p-6 border-b"><div class="flex"><div class="flex-1">
<h1 class="text-2xl">Dr. Jain Sandeep</h1></div>
<div><span>牙醫 Dentists</span><span>牙科普通科</span></div></div>
<div class="flex flex-wrap gap-3 mt-4"><span>男醫生 (Male)</span>
<a href="/district/central-western/">中西區 (Central &amp; Western)</a>
<span>私人執業</span></div></div>
<div class="grid md:grid-cols-2">
<div class="p-6 space-y-5">
<div><h2>電話 Telephone</h2><a href="tel:25222099">25222099</a></div>
<div><h2>地址 Address</h2><p>Room 1914, Melbourne Plaza, 33 Queen's Road Central, Hong Kong</p>
<p>香港中環皇后大道中33號萬邦行1914室</p></div>
<div><h2>傳真 Fax</h2><p>28685336</p></div>
<div><h2>電郵 Email</h2><a href="mailto:drjain@dies電話andpartners.com">drjain@dies電話andpartners.com</a></div>
<div><h2>語言 Languages</h2>  </div>
<div class="flex flex-wrap gap-2">    </div>
</div>
<div class="p-6 space-y-5">
<div><div class="flex items-center gap-2 mb-2"><h2>診症時間 Office Hours</h2>
<span id="open-now-badge"></span></div>
<div class="space-y-1"><div class="flex gap-2"><span>星期一</span><span>0900-1800</span></div>
<div class="flex gap-2"><span>星期六</span><span>0900-1300</span></div></div></div>
<div><h2>資歷 Qualifications</h2><p>香港大學牙醫碩士(牙周病學) MDS (Perio)(HK)</p></div>
</div></div>
<h2>附近同專科醫護人員 Nearby 牙科普通科 Providers</h2>
<div><a href="/providers/dentists/central-western/other-99999/">Dr. Someone Else</a>
<span>女醫生 (Female)</span><span>29998888</span></div>
</div>
<script>(function(){const officeHrRaw = "\\u661f\\u671f\\u4e00\\uff1a\\n 0900-1800 \\n\\n \\u661f\\u671f\\u516d\\uff1a\\n 0900-1300";
const providerPayload = "{\\"id\\":\\"66311\\",\\"nameCN\\":\\"\\u738b\\u5927\\u6587\\",\\"nameEN\\":\\"Dr. Jain Sandeep\\",\\"typeZH\\":\\"\\u7259\\u91ab\\",\\"districtZH\\":\\"\\u4e2d\\u897f\\u5340\\",\\"specialistZH\\":\\"\\u7259\\u79d1\\u666e\\u901a\\u79d1\\",\\"tel\\":\\"25222099\\"}";})();</script>
</body></html>"""


def check_live_markup() -> list[str]:
    """Parse the real-site fixture and assert every field lands correctly."""
    rec = scrape.parse_detail(LIVE_MARKUP, "https://example.test/p/1/", "牙醫")
    expected = {
        # Must be the practitioner, not the site-wide Organization JSON-LD.
        "name": "Dr. Jain Sandeep",
        "name_zh": "王大文",
        "district": "中西區",
        "specialty": "牙科普通科",
        "phone": "25222099",
        "provider_id": "66311",
        # "tel" -> 電話 mangling undone.
        "email": "drjain@diestelandpartners.com",
        # Label stripped from the value.
        "fax": "28685336",
        "address_zh": "香港中環皇后大道中33號萬邦行1914室",
        "hours": "星期一 0900-1800; 星期六 0900-1300",
        "qualifications": "香港大學牙醫碩士(牙周病學) MDS (Perio)(HK)",
        "gender": "男醫生 (Male)",
        "practice": "私人執業",
        # Empty on the page: must stay empty, not absorb the next column or
        # echo its own label back.
        "languages": "",
    }
    failures = [
        f"live markup {key}: got {rec.get(key)!r}, expected {want!r}"
        for key, want in expected.items() if rec.get(key) != want
    ]
    if "Melbourne Plaza" not in rec.get("address"):
        failures.append(f"live markup address: got {rec.get('address')!r}")
    # The nearby-providers list must not leak into this record.
    if "29998888" in " ".join(rec.values.values()):
        failures.append("live markup: nearby provider data leaked into record")
    return failures


def main() -> int:
    with socketserver.TCPServer(("127.0.0.1", 0), Handler) as httpd:
        port = httpd.server_address[1]
        threading.Thread(target=httpd.serve_forever, daemon=True).start()
        scrape.BASE = f"http://127.0.0.1:{port}"

        tmp = Path(tempfile.mkdtemp())
        cache = scrape.Cache(tmp / "cache.sqlite")
        fetcher = scrape.Fetcher(cache, delay=0.0, timeout=10, retries=1)

        categories = scrape.verify_category_slugs(fetcher)
        assert set(categories) == set(CATEGORY_SIZES), categories

        data = {}
        for slug, tab in categories.items():
            data[tab] = scrape.scrape_category(fetcher, slug, tab, None, workers=4)

        failures: list[str] = check_live_markup()

        for slug, tab in categories.items():
            got, want = len(data[tab]), CATEGORY_SIZES[slug]
            if got != want:
                failures.append(f"{tab}: got {got} records, expected {want}")

        required = ("name", "address", "phone", "fax", "email", "district", "specialty")
        for tab, records in data.items():
            for rec in records:
                missing = [k for k in required if not rec.get(k)]
                if missing:
                    failures.append(f"{tab} {rec.url}: missing {missing}")

        # Spot-check exact values across all three markup styles.
        dentists = {r.url: r for r in data["牙醫"]}
        for idx in (0, 1, 2):
            url = f"{scrape.BASE}/doctor/dentists-{idx}/"
            rec = dentists.get(url)
            if not rec:
                failures.append(f"missing profile {url}")
                continue
            checks = {
                "phone": f"2{idx:03d} {1000 + idx}",
                "fax": f"3{idx:03d} {2000 + idx}",
                "email": f"dentists{idx}@example.com",
                "district": DISTRICTS[idx % 3],
                "specialty": SPECIALTIES[idx % 3],
            }
            for key, want in checks.items():
                if rec.get(key) != want:
                    failures.append(
                        f"style {idx} {key}: got {rec.get(key)!r}, expected {want!r}")
            if f"示範道 {idx} 號" not in rec.get("address"):
                failures.append(f"style {idx} address: got {rec.get('address')!r}")

        out = tmp / "out.xlsx"
        scrape.write_workbook(data, out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        expected_tabs = ["說明 Summary", "牙醫", "中醫師", "醫生", "物理治療師"]
        if wb.sheetnames != expected_tabs:
            failures.append(f"tabs: got {wb.sheetnames}, expected {expected_tabs}")
        for slug, tab in categories.items():
            if tab in wb.sheetnames:
                rows = wb[tab].max_row - 1
                if rows != CATEGORY_SIZES[slug]:
                    failures.append(f"{tab} sheet: {rows} rows, expected {CATEGORY_SIZES[slug]}")

        httpd.shutdown()

        if failures:
            print("FAILED:")
            for f in failures:
                print("  -", f)
            return 1
        total = sum(len(r) for r in data.values())
        print(f"PASS - {total} records across {len(data)} tabs, workbook verified")
        print(f"  sample workbook: {out}")
        return 0


if __name__ == "__main__":
    sys.exit(main())
