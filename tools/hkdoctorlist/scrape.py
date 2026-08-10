#!/usr/bin/env python3
"""Scrape hkdoctorlist.com.hk into an Excel workbook with one tab per category.

The site groups practitioners under /type/<slug>/ listing pages that paginate as
/type/<slug>/<page>/, and each row links to a practitioner detail page. This
script walks the listings, fetches every detail page, extracts the contact
fields, and writes 牙醫 / 中醫師 / 醫生 / 物理治療師 tabs.

Every page fetched is cached in a local SQLite file, so re-running is cheap and
the parsing rules can be re-applied without hitting the site again
(`--reparse`). Run `--probe <url>` to dump how a single page is being read.

Usage:
    python scrape.py                       # full scrape -> hkdoctorlist.xlsx
    python scrape.py --limit 25            # 25 practitioners per category
    python scrape.py --reparse             # rebuild the xlsx from the cache
    python scrape.py --probe <url>         # show extraction for one page
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
import threading
import time
import urllib.robotparser
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup

BASE = "https://www.hkdoctorlist.com.hk"
USER_AGENT = (
    "Mozilla/5.0 (compatible; hkdoctorlist-export/1.0; "
    "+contact via repository owner) Python-requests"
)

# Category slug -> Excel tab name. Slugs confirmed from the live site's own
# listing URLs; `physiotherapists` is verified at runtime against /type/ and
# corrected automatically if the site uses a different slug.
CATEGORIES: dict[str, str] = {
    "dentists": "牙醫",
    "chinese-medicine": "中醫師",
    "doctors": "醫生",
    "physiotherapists": "物理治療師",
}

# Path segments that are navigation, never practitioner profiles.
NON_PROFILE_SEGMENTS = {
    "type", "district", "specialist", "specialty", "search", "page", "tag",
    "about", "contact", "privacy", "terms", "disclaimer", "faq", "blog",
    "news", "sitemap", "login", "register", "en", "zh", "tc", "sc", "api",
    "assets", "static", "images", "css", "js",
}


# --------------------------------------------------------------------------
# Field definitions
# --------------------------------------------------------------------------

@dataclass
class Field:
    key: str
    header: str
    labels: tuple[str, ...]


FIELDS: tuple[Field, ...] = (
    Field("name", "姓名 Name", ("姓名", "名稱", "醫生姓名", "中文姓名", "英文姓名", "Name")),
    Field("clinic", "診所名稱 Clinic", ("診所名稱", "診所", "機構名稱", "Clinic", "Practice")),
    Field("address", "地址 Address", ("地址", "診所地址", "辦公地址", "Address")),
    Field("phone", "電話 Phone", ("電話", "聯絡電話", "診所電話", "Tel", "Telephone", "Phone")),
    Field("email", "電郵 Email", ("電郵", "電子郵件", "電郵地址", "Email", "E-mail")),
    Field("fax", "傳真 Fax", ("傳真", "傳真號碼", "Fax")),
    Field("district", "地區 District", ("地區", "區域", "所在地區", "District", "Region", "Location")),
    Field("specialty", "專科 Specialty", ("專科", "註冊專科", "分科", "專業", "科別",
                                          "Specialty", "Speciality", "Specialisation")),
    Field("qualifications", "資格 Qualifications", ("資格", "專業資格", "學歷", "Qualifications")),
    Field("registration", "註冊編號 Registration No.", ("註冊編號", "牌照號碼", "註冊號碼",
                                                        "Registration", "Registration No")),
    Field("languages", "語言 Languages", ("語言", "應診語言", "Languages")),
    Field("hours", "應診時間 Office Hours", ("應診時間", "診症時間", "營業時間",
                                             "Office Hours", "Opening Hours", "Consultation Hours")),
)

FIELD_KEYS = [f.key for f in FIELDS]
EXTRA_HEADERS = ["個人檔案 Profile URL"]

EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")
# HK numbers are 8 digits, often written 1234 5678 / 1234-5678.
PHONE_RE = re.compile(r"(?<!\d)(?:\+?852[\s-]*)?[2-9]\d{3}[\s-]?\d{4}(?!\d)")


# --------------------------------------------------------------------------
# Cache + fetching
# --------------------------------------------------------------------------

class Cache:
    """SQLite-backed page cache so a 10k-page scrape is resumable."""

    def __init__(self, path: Path):
        self.path = path
        self._local = threading.local()
        self._write_lock = threading.Lock()
        with self._connect() as conn:
            conn.execute(
                "CREATE TABLE IF NOT EXISTS pages ("
                " url TEXT PRIMARY KEY, status INTEGER, html TEXT, fetched_at TEXT)"
            )

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.path, timeout=60)
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

    @property
    def conn(self) -> sqlite3.Connection:
        if getattr(self._local, "conn", None) is None:
            self._local.conn = self._connect()
        return self._local.conn

    def get(self, url: str) -> str | None:
        row = self.conn.execute(
            "SELECT html FROM pages WHERE url = ? AND status = 200", (url,)
        ).fetchone()
        return row[0] if row else None

    def put(self, url: str, status: int, html: str) -> None:
        with self._write_lock:
            self.conn.execute(
                "INSERT OR REPLACE INTO pages (url, status, html, fetched_at)"
                " VALUES (?, ?, ?, ?)",
                (url, status, html, datetime.now(timezone.utc).isoformat()),
            )
            self.conn.commit()

    def urls_like(self, pattern: str) -> list[str]:
        rows = self.conn.execute(
            "SELECT url FROM pages WHERE url LIKE ? AND status = 200", (pattern,)
        ).fetchall()
        return [r[0] for r in rows]


class Fetcher:
    def __init__(self, cache: Cache, delay: float, timeout: float, retries: int,
                 obey_robots: bool = True):
        self.cache = cache
        self.delay = delay
        self.timeout = timeout
        self.retries = retries
        self._local = threading.local()
        self._throttle = threading.Lock()
        self._next_ok = 0.0
        self.robots = self._load_robots() if obey_robots else None

    def _load_robots(self):
        rp = urllib.robotparser.RobotFileParser()
        rp.set_url(f"{BASE}/robots.txt")
        try:
            resp = requests.get(f"{BASE}/robots.txt", timeout=15,
                                headers={"User-Agent": USER_AGENT})
            if resp.status_code == 200:
                rp.parse(resp.text.splitlines())
                return rp
        except requests.RequestException:
            pass
        return None  # unreachable robots.txt -> stay polite via rate limit only

    def allowed(self, url: str) -> bool:
        return self.robots.can_fetch(USER_AGENT, url) if self.robots else True

    @property
    def session(self) -> requests.Session:
        if getattr(self._local, "session", None) is None:
            s = requests.Session()
            s.headers.update({
                "User-Agent": USER_AGENT,
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "zh-HK,zh;q=0.9,en;q=0.8",
            })
            self._local.session = s
        return self._local.session

    def _wait_turn(self) -> None:
        if self.delay <= 0:
            return
        with self._throttle:
            now = time.monotonic()
            if now < self._next_ok:
                time.sleep(self._next_ok - now)
                now = time.monotonic()
            self._next_ok = now + self.delay

    def get(self, url: str, force: bool = False) -> str | None:
        if not force:
            cached = self.cache.get(url)
            if cached is not None:
                return cached
        if not self.allowed(url):
            print(f"  robots.txt disallows {url}", file=sys.stderr)
            return None

        backoff = 2.0
        for attempt in range(self.retries + 1):
            self._wait_turn()
            try:
                resp = self.session.get(url, timeout=self.timeout)
            except requests.RequestException as exc:
                if attempt == self.retries:
                    print(f"  fetch failed {url}: {exc}", file=sys.stderr)
                    return None
                time.sleep(backoff)
                backoff *= 2
                continue

            if resp.status_code == 200:
                resp.encoding = resp.encoding or "utf-8"
                self.cache.put(url, 200, resp.text)
                return resp.text
            if resp.status_code == 404:
                self.cache.put(url, 404, "")
                return None
            if resp.status_code in (429, 500, 502, 503, 504) and attempt < self.retries:
                time.sleep(backoff)
                backoff *= 2
                continue
            print(f"  HTTP {resp.status_code} for {url}", file=sys.stderr)
            return None
        return None


# --------------------------------------------------------------------------
# Parsing helpers
# --------------------------------------------------------------------------

def soup_of(html: str) -> BeautifulSoup:
    try:
        return BeautifulSoup(html, "lxml")
    except Exception:
        return BeautifulSoup(html, "html.parser")


def clean(text: str | None) -> str:
    if not text:
        return ""
    text = text.replace("\xa0", " ").replace("　", " ")
    return re.sub(r"\s+", " ", text).strip(" \t\r\n:：-–—|")


def internal_links(soup: BeautifulSoup, page_url: str) -> list[str]:
    out = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
            continue
        full = urljoin(page_url, href)
        parsed = urlparse(full)
        if parsed.netloc and parsed.netloc not in urlparse(BASE).netloc:
            continue
        out.append(parsed._replace(fragment="", query="").geturl())
    return out


def first_segment(url: str) -> str:
    parts = [p for p in urlparse(url).path.split("/") if p]
    return parts[0].lower() if parts else ""


def discover_profile_segment(soup: BeautifulSoup, page_url: str) -> str | None:
    """Infer the URL segment used for practitioner profiles.

    The profile prefix is not hardcoded: whichever non-navigation path segment
    dominates a full listing page's links is treated as the profile segment.
    Only call this on page 1, which is always full - later pages can be short.
    """
    counts: Counter[str] = Counter()
    for link in internal_links(soup, page_url):
        seg = first_segment(link)
        if not seg or seg in NON_PROFILE_SEGMENTS:
            continue
        parts = [p for p in urlparse(link).path.split("/") if p]
        if len(parts) < 2:  # profiles are /<segment>/<identifier>/
            continue
        counts[seg] += 1

    if not counts:
        return None
    best_seg, best_n = counts.most_common(1)[0]
    return best_seg if best_n >= 3 else None


def profile_links(soup: BeautifulSoup, page_url: str, segment: str) -> list[str]:
    """All links under a known profile segment, in page order, deduplicated."""
    seen, out = set(), []
    for link in internal_links(soup, page_url):
        parts = [p for p in urlparse(link).path.split("/") if p]
        if len(parts) >= 2 and parts[0].lower() == segment and link not in seen:
            seen.add(link)
            out.append(link)
    return out


def discover_profile_links(soup: BeautifulSoup, page_url: str) -> list[str]:
    """Convenience wrapper: detect the segment and extract links in one step."""
    segment = discover_profile_segment(soup, page_url)
    return profile_links(soup, page_url, segment) if segment else []


def discover_max_page(soup: BeautifulSoup, page_url: str, slug: str) -> int:
    """Read the highest /type/<slug>/<n>/ page number from pagination links."""
    pattern = re.compile(rf"^/type/{re.escape(slug)}/(\d+)/?$")
    best = 1
    for link in internal_links(soup, page_url):
        m = pattern.match(urlparse(link).path)
        if m:
            best = max(best, int(m.group(1)))
    return best


def label_value_pairs(soup: BeautifulSoup) -> dict[str, str]:
    """Collect label -> value pairs from dl/table/label-ish DOM structures."""
    pairs: dict[str, str] = {}

    def record(label: str, value: str) -> None:
        label, value = clean(label), clean(value)
        if label and value and label not in pairs:
            pairs[label] = value

    for dl in soup.find_all("dl"):
        dts = dl.find_all("dt")
        for dt in dts:
            dd = dt.find_next_sibling("dd")
            if dd:
                record(dt.get_text(" "), dd.get_text(" "))

    for tr in soup.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        if len(cells) >= 2:
            record(cells[0].get_text(" "), cells[1].get_text(" "))

    # Generic "<span>電話</span><span>1234 5678</span>" / "<b>電話:</b> 1234 5678"
    for tag in soup.find_all(["span", "b", "strong", "th", "label", "div", "p", "li"]):
        text = clean(tag.get_text(" "))
        if not text or len(text) > 24:
            continue
        sibling = tag.find_next_sibling()
        if sibling is not None:
            record(text, sibling.get_text(" "))
        elif tag.next_sibling and isinstance(tag.next_sibling, str):
            record(text, tag.next_sibling)

    return pairs


def text_lines(soup: BeautifulSoup) -> list[str]:
    raw = soup.get_text("\n")
    return [clean(line) for line in raw.split("\n") if clean(line)]


def split_label_value(text: str, labels: tuple[str, ...]) -> tuple[bool, str]:
    """Match `text` against a field's labels.

    Returns (matched, inline_value). A label cell sometimes carries the value
    with it ("電郵：a@b.com"), in which case the remainder is returned and
    should be preferred over whatever sits in the adjacent cell.
    """
    candidate = text.strip().strip("*")
    lowered = candidate.lower()
    # Longest label first so "電郵地址" is not consumed by "電郵".
    for label in sorted(labels, key=len, reverse=True):
        lab = label.lower()
        if lowered.rstrip(" \t:：") == lab:
            return True, ""
        if lowered.startswith(lab):
            rest = candidate[len(label):]
            # Require a real separator, so "電話卡" does not match "電話".
            if rest[:1] in (":", "：", "-", "–", "—", " ", "\t"):
                return True, clean(rest)
    return False, ""


def from_jsonld(soup: BeautifulSoup) -> dict[str, str]:
    """Pull fields out of schema.org JSON-LD when the page provides it."""
    found: dict[str, str] = {}

    def walk(node) -> None:
        if isinstance(node, list):
            for item in node:
                walk(item)
            return
        if not isinstance(node, dict):
            return

        def take(key: str, value) -> None:
            value = clean(value if isinstance(value, str) else "")
            if value and key not in found:
                found[key] = value

        take("name", node.get("name"))
        take("phone", node.get("telephone"))
        take("fax", node.get("faxNumber"))
        take("email", node.get("email"))

        spec = node.get("medicalSpecialty") or node.get("jobTitle")
        if isinstance(spec, list):
            spec = ", ".join(str(s) for s in spec if isinstance(s, str))
        take("specialty", spec)

        addr = node.get("address")
        if isinstance(addr, str):
            take("address", addr)
        elif isinstance(addr, dict):
            parts = [addr.get(k) for k in
                     ("streetAddress", "addressLocality", "addressRegion")]
            take("address", ", ".join(clean(p) for p in parts if isinstance(p, str) and p.strip()))
            take("district", addr.get("addressLocality"))

        for value in node.values():
            if isinstance(value, (dict, list)):
                walk(value)

    for script in soup.find_all("script", type="application/ld+json"):
        try:
            walk(json.loads(script.string or ""))
        except (json.JSONDecodeError, TypeError):
            continue
    return found


@dataclass
class Record:
    url: str = ""
    category: str = ""
    values: dict[str, str] = field(default_factory=dict)

    def get(self, key: str) -> str:
        return self.values.get(key, "")


def parse_detail(html: str, url: str, category: str) -> Record:
    soup = soup_of(html)

    values: dict[str, str] = {}

    # 1. Structured data wins when present - read it before scripts are stripped.
    values.update({k: v for k, v in from_jsonld(soup).items() if v})

    for junk in soup.find_all(["script", "style", "noscript"]):
        junk.decompose()

    # 2. Label/value pairs from the DOM.
    pairs = label_value_pairs(soup)
    for fld in FIELDS:
        if values.get(fld.key):
            continue
        for label, value in pairs.items():
            matched, inline = split_label_value(label, fld.labels)
            if not matched:
                continue
            # Prefer a value carried inline with the label over the next cell.
            chosen = inline or clean(value)
            if chosen:
                values[fld.key] = chosen
                break

    # 3. Same-line "label: value" text fallback.
    lines = text_lines(soup)
    for fld in FIELDS:
        if values.get(fld.key):
            continue
        for line in lines:
            m = re.match(
                r"^\s*(?:%s)\s*[:：]\s*(.+)$" % "|".join(re.escape(l) for l in fld.labels),
                line, flags=re.I)
            if m and clean(m.group(1)):
                values[fld.key] = clean(m.group(1))
                break

    # 4. Direct link / pattern fallbacks.
    if not values.get("email"):
        mailto = soup.select_one('a[href^="mailto:"]')
        if mailto:
            values["email"] = clean(mailto["href"].split(":", 1)[1].split("?")[0])
    if not values.get("email"):
        m = EMAIL_RE.search(soup.get_text(" "))
        if m:
            values["email"] = m.group(0)

    if not values.get("phone"):
        tel = soup.select_one('a[href^="tel:"]')
        if tel:
            values["phone"] = clean(tel["href"].split(":", 1)[1])
    if not values.get("phone"):
        m = PHONE_RE.search(soup.get_text(" "))
        if m:
            values["phone"] = clean(m.group(0))

    if not values.get("name"):
        h1 = soup.find(["h1", "h2"])
        if h1:
            values["name"] = clean(h1.get_text(" "))
    if not values.get("name") and soup.title:
        values["name"] = clean(soup.title.get_text().split("|")[0])

    # District/specialty are frequently only expressed as breadcrumb links.
    if not values.get("district"):
        crumb = soup.select_one('a[href*="/district/"]')
        if crumb:
            values["district"] = clean(crumb.get_text(" "))
    if not values.get("specialty"):
        crumb = soup.select_one('a[href*="/specialist/"]')
        if crumb:
            values["specialty"] = clean(crumb.get_text(" "))

    return Record(url=url, category=category, values=values)


# --------------------------------------------------------------------------
# Crawl
# --------------------------------------------------------------------------

def verify_category_slugs(fetcher: Fetcher) -> dict[str, str]:
    """Confirm the /type/<slug>/ slugs against the site's own type index."""
    categories = dict(CATEGORIES)
    html = fetcher.get(f"{BASE}/type/") or fetcher.get(f"{BASE}/")
    if not html:
        return categories

    soup = soup_of(html)
    live: dict[str, str] = {}
    for a in soup.find_all("a", href=True):
        m = re.match(r"^/type/([a-z0-9-]+)/?$", urlparse(urljoin(BASE, a["href"])).path)
        if m:
            live[m.group(1)] = clean(a.get_text(" "))
    if not live:
        return categories

    for slug in list(categories):
        if slug in live:
            continue
        # Slug drifted (e.g. physiotherapists vs physiotherapy) - match on the
        # tab's Chinese name appearing in the live link text.
        want = categories[slug]
        guess = next((s for s, label in live.items() if want and want in label), None)
        if guess:
            print(f"  slug '{slug}' not found; using '{guess}' for {want}")
            categories[guess] = categories.pop(slug)
        else:
            print(f"  warning: no /type/ link found for {want} ('{slug}')", file=sys.stderr)
    return categories


def collect_profile_urls(fetcher: Fetcher, slug: str, limit: int | None) -> list[str]:
    first_url = f"{BASE}/type/{slug}/"
    html = fetcher.get(first_url)
    if not html:
        print(f"  could not load {first_url}", file=sys.stderr)
        return []

    soup = soup_of(html)
    max_page = discover_max_page(soup, first_url, slug)
    # Page 1 is always full, so it is the only safe place to infer the segment.
    segment = discover_profile_segment(soup, first_url)
    if not segment:
        print(f"  could not identify profile links on {first_url}", file=sys.stderr)
        return []
    print(f"  {slug}: {max_page} listing page(s), profiles under /{segment}/")

    seen: set[str] = set()
    ordered: list[str] = []

    def absorb(page_soup: BeautifulSoup, page_url: str) -> int:
        added = 0
        for link in profile_links(page_soup, page_url, segment):
            if link not in seen:
                seen.add(link)
                ordered.append(link)
                added += 1
        return added

    absorb(soup, first_url)

    page = 2
    empty_streak = 0
    while True:
        if limit and len(ordered) >= limit:
            break
        if max_page > 1 and page > max_page:
            break
        page_url = f"{BASE}/type/{slug}/{page}/"
        page_html = fetcher.get(page_url)
        if not page_html:
            break
        # The advertised page count can be stale, so also stop on two dry pages.
        if absorb(soup_of(page_html), page_url) == 0:
            empty_streak += 1
            if empty_streak >= 2:
                break
        else:
            empty_streak = 0
        page += 1

    return ordered[:limit] if limit else ordered


def scrape_category(fetcher: Fetcher, slug: str, tab: str, limit: int | None,
                    workers: int) -> list[Record]:
    print(f"[{tab}] collecting profile links...")
    urls = collect_profile_urls(fetcher, slug, limit)
    print(f"[{tab}] {len(urls)} profiles")
    if not urls:
        return []

    records: list[Record] = []
    done = 0
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(fetcher.get, url): url for url in urls}
        for future in as_completed(futures):
            url = futures[future]
            html = future.result()
            done += 1
            if done % 100 == 0 or done == len(urls):
                print(f"[{tab}] {done}/{len(urls)} fetched")
            if html:
                records.append(parse_detail(html, url, tab))

    order = {url: i for i, url in enumerate(urls)}
    records.sort(key=lambda r: order.get(r.url, 1 << 30))
    return records


def reparse_from_cache(cache: Cache, categories: dict[str, str]) -> dict[str, list[Record]]:
    """Rebuild records from cached HTML without any network access."""
    results: dict[str, list[Record]] = {tab: [] for tab in categories.values()}
    listing_re = re.compile(r"^/(?:type|district|specialist)/")

    profile_urls = [
        u for u in cache.urls_like(f"{BASE}/%")
        if not listing_re.match(urlparse(u).path)
        and len([p for p in urlparse(u).path.split("/") if p]) >= 2
    ]
    if not profile_urls:
        return results

    # Attribute each cached profile to a category by re-reading listing pages.
    owner: dict[str, str] = {}
    for slug, tab in categories.items():
        listings = sorted(cache.urls_like(f"{BASE}/type/{slug}/%"))
        first_url = f"{BASE}/type/{slug}/"
        first_html = cache.get(first_url)
        if not first_html:
            continue
        # Infer the segment from the full first page, then apply it to every
        # listing page - later pages can be too short to detect it on their own.
        segment = discover_profile_segment(soup_of(first_html), first_url)
        if not segment:
            continue
        for listing in [first_url] + [u for u in listings if u != first_url]:
            html = cache.get(listing)
            if html:
                for link in profile_links(soup_of(html), listing, segment):
                    owner.setdefault(link, tab)

    for url in profile_urls:
        tab = owner.get(url)
        if not tab:
            continue
        html = cache.get(url)
        if html:
            results[tab].append(parse_detail(html, url, tab))
    return results


# --------------------------------------------------------------------------
# Excel output
# --------------------------------------------------------------------------

def write_workbook(data: dict[str, list[Record]], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    headers = [f.header for f in FIELDS] + EXTRA_HEADERS

    for tab, records in data.items():
        ws = wb.create_sheet(title=tab[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        for rec in records:
            ws.append([rec.get(k) for k in FIELD_KEYS] + [rec.url])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
        )

        for idx, header in enumerate(headers, start=1):
            longest = len(header)
            for row in range(2, min(ws.max_row, 400) + 1):
                value = ws.cell(row=row, column=idx).value
                if value:
                    longest = max(longest, len(str(value)))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(longest + 2, 12), 60)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = wb.create_sheet(title="說明 Summary", index=0)
    summary.append(["來源 Source", BASE])
    summary.append(["匯出時間 Exported", datetime.now().strftime("%Y-%m-%d %H:%M")])
    summary.append([])
    summary.append(["分頁 Tab", "記錄數 Records"])
    for tab, records in data.items():
        summary.append([tab, len(records)])
    summary.append(["總計 Total", sum(len(r) for r in data.values())])
    for cell in (summary["A1"], summary["A4"], summary["B4"]):
        cell.font = Font(bold=True)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 42

    wb.save(out_path)


def write_jsonl(data: dict[str, list[Record]], path: Path) -> None:
    with path.open("w", encoding="utf-8") as fh:
        for records in data.values():
            for rec in records:
                fh.write(json.dumps(asdict(rec), ensure_ascii=False) + "\n")


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def probe(fetcher: Fetcher, url: str) -> None:
    html = fetcher.get(url, force=True)
    if not html:
        print("could not fetch", url)
        return
    soup = soup_of(html)
    print(f"--- {url} ({len(html)} bytes) ---")
    print("\n[JSON-LD]")
    print(json.dumps(from_jsonld(soup), ensure_ascii=False, indent=2))
    print("\n[label/value pairs]")
    for label, value in list(label_value_pairs(soup).items())[:40]:
        print(f"  {label!r}: {value[:90]!r}")
    print("\n[profile links found]")
    for link in discover_profile_links(soup, url)[:10]:
        print("  ", link)
    print("\n[parsed record]")
    rec = parse_detail(html, url, "probe")
    for fld in FIELDS:
        print(f"  {fld.header:28} {rec.get(fld.key)!r}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out", type=Path, default=Path("hkdoctorlist.xlsx"))
    ap.add_argument("--cache", type=Path, default=Path("hkdoctorlist-cache.sqlite"))
    ap.add_argument("--jsonl", type=Path, help="also dump raw records as JSONL")
    ap.add_argument("--limit", type=int, help="max practitioners per category (testing)")
    ap.add_argument("--workers", type=int, default=4, help="concurrent fetches")
    ap.add_argument("--delay", type=float, default=0.4, help="seconds between requests")
    ap.add_argument("--timeout", type=float, default=30.0)
    ap.add_argument("--retries", type=int, default=3)
    ap.add_argument("--ignore-robots", action="store_true")
    ap.add_argument("--reparse", action="store_true",
                    help="rebuild the workbook from cache, no network")
    ap.add_argument("--probe", metavar="URL", help="dump extraction for one page")
    args = ap.parse_args()

    cache = Cache(args.cache)
    fetcher = Fetcher(cache, args.delay, args.timeout, args.retries,
                      obey_robots=not args.ignore_robots)

    if args.probe:
        probe(fetcher, args.probe)
        return 0

    if args.reparse:
        print("reparsing from cache (no network)...")
        data = reparse_from_cache(cache, CATEGORIES)
    else:
        print("verifying category slugs...")
        categories = verify_category_slugs(fetcher)
        data = {}
        for slug, tab in categories.items():
            data[tab] = scrape_category(fetcher, slug, tab, args.limit, args.workers)

    total = sum(len(r) for r in data.values())
    for tab, records in data.items():
        filled = sum(1 for r in records if r.get("phone"))
        print(f"  {tab}: {len(records)} records ({filled} with phone)")
    if total == 0:
        print("no records scraped - nothing written", file=sys.stderr)
        return 1

    write_workbook(data, args.out)
    print(f"wrote {args.out} ({total} records)")
    if args.jsonl:
        write_jsonl(data, args.jsonl)
        print(f"wrote {args.jsonl}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
